from re import error
from datetime import date, datetime
import datetime, requests, json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied, RequestDataTooBig
from django.db.models import QuerySet, Count, Q
from django.http.request import RawPostDataException
from django.http import HttpResponse


from rest_framework import generics, viewsets, status
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.permissions import IsAdminUser, IsAuthenticated, AllowAny
from rest_framework.serializers import Serializer
from rest_framework.viewsets import ModelViewSet, ViewSet

from .models import *
from .serializers import *
from .permissions import *
from .exceptions import CustomError



headers = {
    'Authorization': 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1aWQiOiI3MzcwZDc4Ni1lN2Y1LTQzNmYtYjYwOC0wYTc2NDRmZGI4ZTciLCJyb2wiOiJ1c2VyIiwiYXVkIjoidzRGNDV2cDVicGxldEFGZE5pWnhVRUV6cWFTemZ3SzAiLCJpYXQiOjE2MTI5NzQ5MDcsImlzcyI6InNsaWNrcG9zIn0.i7rTScUlAmZPK_jcoD-wQsP5OitUBFsEjjiRoQmcYaw'
}


# ------------------------- Begin State City Country ------------------------ #

class StateListAPIView(generics.ListAPIView):
    serializer_class = StateSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return State.objects.filter(country_id=101).order_by('-id')


class CityListAPIView(generics.ListAPIView):
    serializer_class = CitySerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return City.objects.filter(state_id=self.kwargs['state_id']).order_by('-id')


# ------------------------- End State City Country ------------------------ #

class AuthLoginAPIView(generics.CreateAPIView):
    serializer_class = UserTokenSerializer

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        token, created = Token.objects.get_or_create(user=user)
        response = BaseUserSerializer(user).data
        response['token'] = token.key
        return Response(response)


class UserAPIView(viewsets.ModelViewSet):
    queryset = BaseUser.objects.filter(is_active=True)
    serializer_class = UserSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = BaseUser.objects.filter(pk=kwargs['pk']).update(is_active=False)
        return Response({'message': 'BaseUser deleted sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class AuthVerifyAPIView(generics.RetrieveAPIView):
    serializer_class = BaseUserSerializer

    def get_object(self):
        return self.request.user


class RoleListAPIView(generics.ListAPIView):
    queryset = EmployeeRole.objects.filter(status=True, delete=False)
    serializer_class = RoleSerializer
    permission_classes = (IsAuthenticated,)


class BranchAPIViewset(viewsets.ModelViewSet):
    queryset = Branch.objects.filter(delete=False).order_by('-pk')
    serializer_class = BranchSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        instance = serializer.save()
        codes = [1, 2, 3, 4]
        for role in codes:
            branch_employee = BranchEmployeeIncentive.objects.create(branch=instance, employee_role=EmployeeRole.objects.get(code=role))
            for department in codes:
                BranchDepartmentIncentive.objects.create(department=BranchProductDepartment.objects.get(code=department), role=branch_employee, incentive=0)

    def destroy(self, request, *args, **kwargs):
        destroy = Branch.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'Branch deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class SubBranchCreate(generics.UpdateAPIView):
    queryset = SubBranch.objects.filter(delete=False)
    serializer_class = SubBranchSerializer
    permission_classes = (IsAuthenticated,)

    def partial_update(self, request, pk):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        sub_branch_data = SubBranch.objects.create(name=serializer.validated_data['name'])
        branch = Branch.objects.get(pk=pk, delete=False)
        branch.sub_branch.add(sub_branch_data)
        return Response({'message': 'SubBranch created sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class SubBranchUpdateApiView(generics.UpdateAPIView):
    queryset = SubBranch.objects.filter(delete=False)
    serializer_class = SubBranchUpdateSerializer
    permission_classes = (IsAuthenticated,)


class SubBranchRetUpdDelAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SubBranch.objects.filter(delete=False)
    serializer_class = SubBranchSerializer
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request, pk):
        return Response(SubBranchSerializer(Branch.objects.get(pk=pk).sub_branch.filter(delete=False), many=True).data)

    def partial_update(self, request, pk):
        instance = self.get_object()
        destroy = SubBranch.objects.filter(pk=pk).update(status=not instance.status)
        return Response({'message': 'branch status update successfully'}, status=status.HTTP_204_NO_CONTENT)

    def destroy(self, request, pk):
        destroy = SubBranch.objects.filter(pk=pk).update(delete=True)
        return Response({'message': 'branch deleted sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class UserSalaryAPIViewset(viewsets.ModelViewSet):
    queryset = UserSalary.objects.filter(status=True, delete=False)
    serializer_class = UserSalarySerializer
    permission_classes = (IsAuthenticated,)


class UserSalaryList(generics.ListAPIView):
    serializer_class = UserSalarySerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        data = UserSalary.objects.filter(user=self.kwargs['user_id']).latest('created').pk
        return UserSalary.objects.filter(pk=data)


class UserSalaryReport(generics.RetrieveAPIView):
    serializer_class = UserSalaryReportSerializer
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request, branch_id):
        try:
            year = datetime.datetime.strptime(request.query_params.get('year'), '%Y')
            month = datetime.datetime.strptime(request.query_params.get('month'), '%m')
            year = year.year
            month = month.month
        except:
            today = datetime.date.today()
            first = today.replace(day=1)
            lastMonth = first - datetime.timedelta(days=1)
            year = lastMonth.year
            month = lastMonth.month

        context = {'branch_id': branch_id, 'month': month, 'year': year}

        if branch_id == 0:
            queryset = UserSalaryPerDay.objects.filter(date__year=year, date__month=month, delete=False)
            query = BaseUser.objects.filter(is_active=True, is_superuser=False, is_staff=False)
        else:
            queryset = UserSalaryPerDay.objects.filter(user__branch__pk=branch_id, date__year=year, date__month=month,delete=False)
            query = BaseUser.objects.filter(branch__pk=branch_id, is_active=True, is_superuser=False, is_staff=False)

        total_salary = queryset.aggregate(total_salary_price=Coalesce(Sum('salary'), 0))
        total_ot = queryset.aggregate(overall_ot_price=Coalesce(Sum('ot_salary'), 0))
        sales_data = UserSalaryReportSerializer(query, context=context, many=True).data

        return Response({
            'month': month,
            'year': year,
            'total_salary': total_salary['total_salary_price'],
            'total_ot': total_ot['overall_ot_price'],
            'data': sales_data
        })


class UserSalaryAttendanceReport(generics.RetrieveAPIView):
    serializer_class = UserSalaryAttendanceReportSerializer
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request, user_id):
        try:
            year = datetime.datetime.strptime(request.query_params.get('year'), '%Y')
            month = datetime.datetime.strptime(request.query_params.get('month'), '%m')
            year = year.year
            month = month.month
        except:
            today = datetime.date.today()
            first = today.replace(day=1)
            lastMonth = first - datetime.timedelta(days=1)
            year = lastMonth.year
            month = lastMonth.month

        context = {'user_id': user_id, 'month': month, 'year': year}

        user = BaseUser.objects.get(pk=user_id)
        queryset = UserSalaryPerDay.objects.filter(user=user_id, date__year=year, date__month=month, status=True, delete=False).order_by('date').distinct('date')
        attendance_data = UserSalaryAttendanceReportSerializer(queryset, context=context, many=True).data

        return Response({
            'user': user.pk,
            'user_name': user.get_full_name(),
            'month': month,
            'year': year,
            'attendance_data': attendance_data
        })


class UserSalaryAttendanceListAPIView(generics.ListAPIView):
    serializer_class = UserSalaryAttendanceListSerializer
    permission_classes = (IsAuthenticated,)

    def get_serializer_context(self):
        try:
            date = datetime.datetime.strptime(self.request.query_params.get('date'), '%Y-%m-%d')
        except:
            today = datetime.date.today()
            date = today - datetime.timedelta(days=1)

        return {
            'request': self.request,
            'format': self.format_kwarg,
            'view': self,
            'date': date
        }

    def get_queryset(self):
        if self.kwargs['branch_id'] == 0:
            data = BaseUser.objects.filter(is_employee=True)
        else:
            data = BaseUser.objects.filter(is_employee=True, branch__pk=self.kwargs['branch_id'])
        return data


class PreviousDateAPIView(generics.RetrieveAPIView):
    serializer_class = UserSalaryAttendanceReportSerializer
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request):
        today = datetime.date.today()
        date = today - datetime.timedelta(days=1)
        return Response({
            'date': date
        })


class UserInAttendanceCreateAPIView(generics.CreateAPIView):
    serializer_class = UserAttendanceInSerializer
    permission_classes = (IsAuthenticated,)


class UserOutAttendanceUpdateAPIView(generics.UpdateAPIView):
    queryset = UserAttendance.objects.filter(delete=False)
    serializer_class = UserAttendanceOutSerializer
    permission_classes = (IsAuthenticated,)


class UserPunchUpdateAPIView(generics.UpdateAPIView):
    queryset = UserAttendance.objects.filter(delete=False)
    serializer_class = UserPunchUpdateSerializer
    permission_classes = (IsAuthenticated,)

    def partial_update(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        for data in serializer.validated_data['data']:
            attendance_data = UserAttendance.objects.get(pk=data['id'])
            attendance_data.start = data['start']
            attendance_data.stop = data['stop']
            attendance_data.save()
        return Response({'message': 'Updated Successfully'}, status=status.HTTP_202_ACCEPTED)


class UserInAttendanceBreakCreateAPIView(generics.CreateAPIView):
    queryset = UserAttendanceBreak.objects.filter(delete=False)
    serializer_class = UserAttendanceBreakInSerializer
    permission_classes = (IsAuthenticated,)

    def post(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        if not UserAttendance.objects.filter(user=serializer.validated_data['user'], date=serializer.validated_data['date']).exists():
            return Response({'message': 'You have to create attendance to use this option'}, status=status.HTTP_202_ACCEPTED)
        else:
            return self.create(request, *args, **kwargs)


class UserOutAttendanceBreakUpdateAPIView(generics.UpdateAPIView):
    queryset = UserAttendanceBreak.objects.filter(delete=False)
    serializer_class = UserAttendanceBreakOutSerializer
    permission_classes = (IsAuthenticated,)


class GSTListAPIView(generics.ListAPIView):
    queryset = GST.objects.filter(delete=False, status=True)
    serializer_class = GSTSerializer
    permission_classes = (IsAuthenticated,)


class UnitListAPIView(generics.ListAPIView):
    queryset = Unit.objects.filter(delete=False, status=True)
    serializer_class = UnitSerializer
    # permission_classes = (IsSuperOrAdminUser,)
    permission_classes = (IsAuthenticated,)


class UnitUpdateAPIView(generics.UpdateAPIView):
    queryset = Unit.objects.filter(delete=False, status=True)
    serializer_class = UnitSerializer
    permission_classes = (IsAuthenticated,)


class RoleUpdateAPIView(generics.UpdateAPIView):
    queryset = EmployeeRole.objects.filter(delete=False)
    serializer_class = RoleSerializer
    permission_classes = (IsAuthenticated,)


class BranchProductClassificationViewset(viewsets.ModelViewSet):
    queryset = BranchProductClassification.objects.exclude(delete=True, status=False)
    serializer_class = BranchProductClassificationSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = BranchProductClassification.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'product category deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class ProductClassificationListAPIView(generics.ListAPIView):
    queryset = BranchProductClassification.objects.filter(status=True, delete=False)
    serializer_class = BranchProductClassificationSerializer
    permission_classes = (IsAuthenticated,)


class BranchProductDepartmentViewset(viewsets.ModelViewSet):
    queryset = BranchProductDepartment.objects.exclude(delete=True, status=False)
    serializer_class = BranchProductDepartmentSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = BranchProductDepartment.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'product type deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class ProductDepartmentListAPIView(generics.ListAPIView):
    queryset = BranchProductDepartment.objects.filter(status=True, delete=False)
    serializer_class = BranchProductDepartmentSerializer
    permission_classes = (IsAuthenticated,)


class UserRoleListAPIView(generics.ListAPIView):
    serializer_class = RoleSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return BaseUser.objects.get(pk=self.kwargs['pk']).employee_role.exclude(code=1)


class ProductRecipeItemViewset(viewsets.ModelViewSet):
    queryset = ProductRecipeItem.objects.filter(delete=False)
    serializer_class = ProductRecipeItemSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = ProductRecipeItem.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'recipe item deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class ProductListAPI(generics.ListAPIView):
    serializer_class = ProductSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        if self.kwargs['classification'] == 0:
            product = Product.objects.filter(status=True, delete=False).order_by('-id')
        else:
            product = Product.objects.filter(classification__code=self.kwargs['classification'], status=True,
                                             delete=False).order_by('-id')
        return product


class ProductCreate(generics.CreateAPIView):
    serializer_class = ProductSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(
            classification=BranchProductClassification.objects.get(code=serializer.validated_data['classification']))


class ProductRetriveUpdateDelete(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.exclude(delete=True, status=False).order_by('-id')
    serializer_class = ProductSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = Product.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'Branch product deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class WrongBillAPIView(viewsets.ModelViewSet):
    serializer_class = WrongBillSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return WrongBill.objects.filter(branch=self.request.user.branch, delete=False, status=True)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch, billed_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        destroy = WrongBill.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'wrong bill deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class BranchSpecificWrongBillAPIView(generics.ListAPIView):
    queryset = WrongBill.objects.filter(delete=False, status=True)
    serializer_class = WrongBillSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return WrongBill.objects.filter(date__date=self.kwargs['date'], branch=self.request.user.branch,delete=False, status=True)


class FreeBillCustomerListAPIView(generics.ListAPIView):
    queryset = FreeBillCustomer.objects.exclude(delete=True, status=False)
    serializer_class = FreeBillCustomerSerializer
    permission_classes = (IsAuthenticated,)


class FreeBillAPIView(viewsets.ModelViewSet):
    queryset = FreeBill.objects.filter(delete=False, status=True)
    serializer_class = FreeBillSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return FreeBill.objects.filter(branch=self.request.user.branch, delete=False, status=True)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch, billed_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        destroy = FreeBill.objects.filter(pk=kwargs['pk']).update(status=False)
        return Response({'message': 'wrong bill deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class BranchSpecificFreeBillAPIView(generics.ListAPIView):
    queryset = FreeBill.objects.filter(delete=False, status=True)
    serializer_class = FreeBillSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return FreeBill.objects.filter(date__date=self.kwargs['date'], branch=self.request.user.branch, delete=False, status=True)


class ComplaintListCreateAPIView(generics.ListCreateAPIView):
    queryset = Complaint.objects.exclude(delete=True, status=False)
    serializer_class = ComplaintSerializer
    # permission_classes = (IsSuperOrAdminUser,)
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(complainted_by=self.request.user, status=ComplaintStatus.objects.get(code=1))


class BulkOrderItemListAPIView(generics.ListAPIView):
    queryset = BulkOrderItem.objects.exclude(delete=True, status=False)
    serializer_class = BulkOrderItemSerializer
    permission_classes = (IsAuthenticated,)


class BulkOrderListCreateAPIView(generics.ListCreateAPIView):
    queryset = BulkOrder.objects.exclude(delete=True, status=False)
    serializer_class = BulkOrderSerializer
    permission_classes = (IsAuthenticated,)

    def get_serializer_context(self):
        return {
            'request': self.request,
            'format': self.format_kwarg,
            'view': self,
        }

    def get_queryset(self):
        return BulkOrder.objects.filter(branch=self.request.user.branch, delete=False, status=True)


class OrderStatusListAPIView(generics.ListAPIView):
    queryset = OrderStatus.objects.exclude(delete=True, status=False)
    serializer_class = OrderStatusSerializer
    permission_classes = (IsAuthenticated,)


# class CustomerListAPIView(generics.ListAPIView):
#     queryset = Customer.objects.exclude(delete=True, status=False)
#     serializer_class = CustomerSerializer
#     # permission_classes = (IsSuperOrAdminUser,)


class BranchProductMappingList(generics.ListAPIView):
    serializer_class = ProductBranchMappingSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, classification):
        try:
            product_mapping = ProductBranchMapping.objects.get(branch=self.request.user.branch).product.values_list('pk', flat=True)
        except:
            product_mapping = [0]
        if self.kwargs['classification'] == 0:
            product = Product.objects.filter(classification__in=[2, 3, 5], status=True, delete=False).exclude(pk__in=product_mapping).order_by('-id')
        else:
            product = Product.objects.filter(classification__code=self.kwargs['classification'], status=True, delete=False).exclude(pk__in=product_mapping).order_by('-id')
        return Response(ProductSerializer(product, many=True).data, status=status.HTTP_201_CREATED)


class BranchProductMappingCreate(generics.CreateAPIView):
    serializer_class = ProductBranchMappingSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request):
        product_mapping, created = ProductBranchMapping.objects.get_or_create(branch=self.request.user.branch)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        product_mapping.product.add(*serializer.validated_data['product'])
        product_mapping.save()
        serializer_data = ProductBranchMappingSerializer(product_mapping).data
        return Response(serializer_data, status=status.HTTP_201_CREATED)


class BranchProductMappingDelete(generics.DestroyAPIView):
    serializer_class = ProductBranchMappingSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, pk):
        product_mapping = ProductBranchMapping.objects.get(branch=self.request.user.branch)
        product_mapping.product.remove(pk)
        product_mapping.save()
        return Response({'message': 'product removed deleted sucessfully'}, status=status.HTTP_200_OK)


class ComplaintStatusViewSet(viewsets.ModelViewSet):
    queryset = ComplaintStatus.objects.exclude(delete=True, status=False)
    serializer_class = ComplaintStatusSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = ComplaintStatus.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'ComplaintStatus deleted sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class ComplaintTypeViewSet(viewsets.ModelViewSet):
    queryset = ComplaintType.objects.exclude(delete=True, status=False)
    serializer_class = ComplaintTypeSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = ComplaintType.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'ComplaintType deleted sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class BranchExpensesViewSet(viewsets.ModelViewSet):
    queryset = BranchExpenses.objects.exclude(delete=True)
    serializer_class = BranchExpensesSerializer
    # permission_classes = (IsSuperOrAdminUser,)
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = BranchExpenses.objects.filter(pk=kwargs['pk']).update(delete=True)
        return Response({'message': 'BranchExpensesdeleted sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class PaymentModeListAPI(generics.ListAPIView):
    queryset = PaymentMode.objects.exclude(delete=True)
    serializer_class = PaymentModeSerializer
    permission_classes = (IsAuthenticated,)


class ProductForMappingList(generics.ListAPIView):
    serializer_class = ProductSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        try:
            product_mapping = ProductBranchMapping.objects.get(
                branch=Branch.objects.get(pk=self.kwargs['branch_id'], delete=False, status=True))
            return Product.objects.exclude(pk__in=product_mapping.product.values_list('pk', flat=True)).exclude(delete=True)
        except ProductBranchMapping.DoesNotExist:
            raise ValidationError({'message': 'Data does not exist'})


class UserAttendanceListAPIView(generics.ListAPIView):
    queryset = BaseUser.objects.filter(is_active=True)
    serializer_class = UserAttendanceListSerializer
    permission_classes = (IsAuthenticated,)

    def get_serializer_context(self):
        return {
            'request': self.request,
            'format': self.format_kwarg,
            'view': self,
            'date': self.kwargs['date']
        }

    def get_queryset(self):
        salary_data = UserSalary.objects.filter(date__lte=self.kwargs['date'], status=True, delete=False).values_list('user__pk', flat=True).distinct()
        return BaseUser.objects.filter(pk__in=salary_data, is_superuser=False, branch=self.request.user.branch, is_active=True)


class AttendanceUserListAPIView(generics.ListAPIView):
    serializer_class = UserListSerializer
    permission_classes = (IsAuthenticated,)

    def get_serializer_context(self):
        return {
            'request': self.request,
            'format': self.format_kwarg,
            'view': self,
            'date': self.kwargs['date']
        }

    def get_queryset(self):
        user = BaseUser.objects.filter(is_superuser=False, date_of_resignation__date__gte=self.kwargs['date'], date_of_joining__date__lte=self.kwargs['date'])
        branch_user = user.filter(branch__pk=self.kwargs['pk'])
        return branch_user


class BranchSpecificUserListAPIView(generics.ListAPIView):
    queryset = BaseUser.objects.filter(is_active=True)
    serializer_class = UserSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return BaseUser.objects.filter(branch=self.kwargs["branch_id"])


class ElectricBillAPIView(viewsets.ModelViewSet):
    queryset = ElectricBill.objects.exclude(delete=True, status=False)
    serializer_class = ElectricBillSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = ElectricBill.objects.filter(pk=kwargs['pk']).update(delete=True, status=False)
        return Response({'message': 'electric_bill deleted sucessfully'}, status=status.HTTP_204_NO_CONTENT)


class ElectricBillMeterList(generics.ListAPIView):
    serializer_class = ElectricBillMeterSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, date):
        query = EBMeter.objects.filter(branch=self.request.user.branch, delete=False, status=True).order_by('-id')
        return Response(ElectricBillMeterSerializer(query, context={'branch': self.request.user.branch, 'date': self.kwargs['date']}, many=True).data)


class BranchElectricBillMeterList(generics.ListAPIView):
    serializer_class = EbMeterSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, branch):
        query = EBMeter.objects.filter(branch=branch, delete=False, status=True).order_by('-id')
        return Response(EbMeterSerializer(query, many=True).data)


class EbMeterCreateAPIView(generics.CreateAPIView):
    queryset = EBMeter.objects.filter(delete=False, status=True)
    serializer_class = EbMeterSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=Branch.objects.get(pk=self.kwargs['branch']))


class EbMeterListAPIView(generics.ListAPIView):
    queryset = EBMeter.objects.filter(delete=False, status=True)
    serializer_class = EbMeterSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return EBMeter.objects.filter(branch=Branch.objects.get(pk=self.kwargs['branch']), delete=False, status=True)


class EbMeterUpdateAPIView(generics.UpdateAPIView):
    queryset = EBMeter.objects.filter(delete=False, status=True)
    serializer_class = EbMeterSerializer
    permission_classes = (IsAuthenticated,)


class EbMeterDestroyAPIView(generics.DestroyAPIView):
    queryset = EBMeter.objects.filter(delete=False, status=True)
    serializer_class = EbMeterSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, pk):
        destroy = EBMeter.objects.filter(pk=pk).update(status=False, delete=True)
        return Response({'message': 'eb meter deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class SubBranchlistAPIView(generics.ListAPIView):
    serializer_class = SubBranchListSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, branch):
        sub_branch = Branch.objects.get(pk=branch, delete=False, status=True).sub_branch.order_by('-id')
        return Response(SubBranchListSerializer(sub_branch, many=True).data)


class ProductPricingBatchAPIView(viewsets.ModelViewSet):
    queryset = ProductPricingBatch.objects.filter(delete=False, status=True)
    serializer_class = ProductPricingBatchSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return ProductPricingBatch.objects.filter(branch=self.request.user.branch, delete=False, status=True)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = ProductPricingBatch.objects.filter(pk=kwargs['pk']).update(delete=True, status=False)
        return Response({'message': 'product batch deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class ProductInventoryAPIView(viewsets.ModelViewSet):
    queryset = ProductInventory.objects.filter(delete=False, status=True)
    serializer_class = ProductInventorySerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return ProductInventory.objects.filter(branch=self.request.user.branch, delete=False, status=True)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = ProductInventory.objects.filter(pk=kwargs['pk']).update(delete=True, status=False)
        return Response({'message': 'product batch deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class AttendanceBulkCreateAPIView(generics.CreateAPIView):
    queryset = UserAttendance.objects.exclude(delete=True).order_by('pk')
    serializer_class = BulkAttendanceSerializer
    attendance_serializer_class = BulkAttendanceSerializer
    break_serializer_class = BulkBreakTimeSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request, *args, **kwargs):
        for data in request.data:
            user = BaseUser.objects.get(pk=data['id'])
            for user_attendance_data in data['user_attendance']:
                serializer = self.attendance_serializer_class(data=user_attendance_data, context={'request': request, 'user': user})
                serializer.is_valid(raise_exception=True)
                serializer.save()
            for break_time_data in data['break_time']:
                serializer = self.break_serializer_class(data=break_time_data, context={'request': request, 'user': user})
                serializer.is_valid(raise_exception=True)
                serializer.save()
        return Response({'message': 'Data Saved!'})


class AttendanceReportListAPIView(generics.RetrieveAPIView):
    queryset = UserAttendance.objects.exclude(delete=True).order_by('pk')
    permission_classes = (IsAuthenticated,)


    def retrieve(self, request, pk):
        start = datetime.datetime.strptime(request.query_params.get('start'), '%Y-%m-%d')
        stop = datetime.datetime.strptime(request.query_params.get('stop'), '%Y-%m-%d')
        user = BaseUser.objects.get(pk=pk, is_active=True)

        attendance_data = AttendanceSerializer(UserAttendance.objects.filter(user__pk=pk).filter(date__range=[start, stop], status=True, delete=False).order_by('-date'), many=True).data
        break_data = AttendanceBreakSerializer(UserAttendanceBreak.objects.filter(user__pk=pk).filter(date__range=[start, stop], status=True, delete=False).order_by('-date'), many=True).data

        return Response({
            'username': user.first_name,
            'attendance': attendance_data,
            'break': break_data,
        })


class UserListAPIView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        if self.kwargs['branch_id'] == 0:
            user = BaseUser.objects.filter(is_employee=True, is_active=True)
        else:
            user = BaseUser.objects.filter(branch=self.kwargs['branch_id'], is_active=True, is_employee=True)
        data = user.exclude(Q(is_superuser=True) | Q(is_staff=True))
        return data


class AdminUserListAPIView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return BaseUser.objects.filter(Q(is_superuser=True) | Q(is_staff=True)).filter(is_employee=False, is_active=True)


class BranchIncentiveListAPIView(generics.ListAPIView):
    queryset = BranchEmployeeIncentive.objects.filter(delete=False)
    serializer_class = BranchEmployeeIncentiveSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return BranchEmployeeIncentive.objects.filter(branch=self.kwargs['pk'], status=True, delete=False)


class BranchIncentiveUpdateAPIView(generics.UpdateAPIView):
    serializer_class = BranchDepartmentIncentiveUpdateSerializer
    permission_classes = (IsAuthenticated,)

    def update(self, request, *args, **kwargs):
        serializer = self.serializer_class(data=request.data, many=True)
        serializer.is_valid(raise_exception=True)
        for data in serializer.validated_data:
            formated_data = dict(data)
            BranchDepartmentIncentive.objects.filter(pk=formated_data['id'].pk).update(incentive=formated_data['incentive'])
        return Response(BranchEmployeeIncentiveSerializer(BranchEmployeeIncentive.objects.filter(branch=self.kwargs['pk'], status=True, delete=False),many=True).data)


class VendorAPIView(viewsets.ModelViewSet):
    queryset = Vendor.objects.filter(delete=False, status=True).order_by('-id')
    serializer_class = VendorSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = Vendor.objects.filter(pk=kwargs['pk']).update(delete=True, status=False)
        return Response({'message': 'vendor deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class BranchProductList(generics.ListAPIView):
    serializer_class = ProductSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, classification):
        try:
            product_mapping = ProductBranchMapping.objects.get(branch=self.request.user.branch).product.order_by('-id')
            if self.kwargs['classification'] == 0:
                products = product_mapping.filter(classification__code__in=[2, 3, 4])
            else:
                products = product_mapping.filter(classification__code=self.kwargs['classification'])
            return Response(ProductSerializer(products, many=True).data)
        except ProductBranchMapping.DoesNotExist:
            return Response([], status=status.HTTP_200_OK)


class InventoryRawProductList(generics.ListAPIView):
    serializer_class = DailySheetInventoryListSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, date):
        user = Branch.objects.get(pk=self.request.user.branch.pk, delete=False, status=True)
        return Response(DailySheetInventoryListSerializer(Branch.objects.get(pk=self.request.user.branch.pk), context={'branch': self.request.user.branch.pk, 'date': date}).data)


class ProductInstockCountList(generics.RetrieveAPIView):
    serializer_class = ProductInstockListSerializer
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request, *args, **kwargs):
        try:
            serializer = self.serializer_class(ProductInventory.objects.get(branch=self.request.user.branch, product__pk=self.kwargs['product']))
            return Response(serializer.data)
        except ProductInventory.DoesNotExist:
            return Response({'message': 'This Product Has No Existing Data'}, status=status.HTTP_400_BAD_REQUEST)


class ProductInventoryControlList(generics.ListAPIView):
    serializer_class = ProductInventoryControlSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, date, classification):
        query = ProductBranchMapping.objects.get(branch=self.request.user.branch).product.order_by('-id')
        products = query.filter(classification__code=classification).order_by('-id')
        return Response(ProductInventoryControlSerializer(products, context={'branch': self.request.user.branch.pk, 'date': date}, many=True).data)


class ProductInventoryControlListByBranch(generics.ListAPIView):
    serializer_class = ProductInventoryControlSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, date, classification, branch):
        try:
            query = ProductBranchMapping.objects.get(branch=branch).product.order_by('-id')
            products = query.filter(classification__code=classification).order_by('-id')
            return Response(ProductInventoryControlSerializer(products, context={'branch': branch, 'date': date}, many=True).data)
        except ProductBranchMapping.DoesNotExist:
            return Response([], status=status.HTTP_200_OK)


class ProductInventoryControlCreate(generics.CreateAPIView):
    serializer_class = ProductInventoryControlCreateSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request, date):
        for inventory_data in request.data:
            if inventory_data['is_editable']:
                if int(inventory_data['closing_stock']) > int(inventory_data['on_hand']):
                    return Response({'message': 'Closing stock is greater then inventory stock'}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    serializer = self.serializer_class(data=inventory_data, context={'branch': self.request.user.branch.pk, 'date': date})
                    serializer.is_valid(raise_exception=True)
                    serializer.save()
            else:
                continue
        return Response({'message': 'Data Saved!'})


class BranchProductInventoryCreate(generics.CreateAPIView):
    serializer_class = BranchProductInventoryCreateSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request):
        for inventory_data in request.data:
            serializer = self.serializer_class(data=inventory_data, context={'branch': self.request.user.branch.pk})
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return Response({'message': 'Data Saved!'})


class BranchProductInventoryList(generics.ListAPIView):
    serializer_class = BranchProductInventoryListSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        start = datetime.datetime.strptime(self.request.query_params.get('start'), '%Y-%m-%d')
        stop = datetime.datetime.strptime(self.request.query_params.get('stop'), '%Y-%m-%d')
        if self.kwargs['pk'] == 1:
            query = ProductPricingBatch.objects.filter(date__date__range=[start, stop], branch=self.request.user.branch.pk, product__classification__code__in=[2, 3], delete=False, status=True).order_by('-id')
        elif self.kwargs['pk'] == 2:
            query = ProductPricingBatch.objects.filter(date__date__range=[start, stop], branch=self.request.user.branch.pk, product__classification__code=4, delete=False, status=True).order_by('-id')
        return query


class OilConsumptionList(generics.ListAPIView):
    serializer_class = OilConsumptionSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return OilConsumption.objects.filter(branch=self.request.user.branch.pk, date__date=self.kwargs['date'], delete=False, status=True).order_by('-id')


class OilConsumptionCreate(generics.CreateAPIView):
    serializer_class = OilConsumptionSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)


class OilConsumptionUpdate(generics.UpdateAPIView):
    queryset = OilConsumption.objects.exclude(delete=True, status=False)
    serializer_class = OilConsumptionSerializer
    permission_classes = (IsAuthenticated,)


class FoodWastageAPIView(viewsets.ModelViewSet):
    queryset = FoodWastage.objects.filter(delete=False, status=True)
    serializer_class = FoodWastageSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return FoodWastage.objects.filter(branch=self.request.user.branch, delete=False, status=True)

    def perform_create(self, serializer):
        serializer.save(branch=serializer.validated_data['wasted_by'].branch)

    def destroy(self, request, *args, **kwargs):
        destroy = FoodWastage.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'food wastage deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class FoodWastageList(generics.ListAPIView):
    serializer_class = FoodWastageSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return FoodWastage.objects.filter(date__date=self.kwargs['date'], branch=self.request.user.branch, delete=False, status=True)


class RawOperationalProductList(generics.ListAPIView):
    serializer_class = RawOperationalProductListSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return ProductBranchMapping.objects.get(branch=self.request.user.branch).product.filter(classification__code__in=[2, 3], status=True, delete=False)


class BranchSpecificUserListAPIView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        user = BaseUser.objects.filter(branch=self.request.user.branch, is_active=True, is_employee=True)
        data = user.exclude(Q(is_superuser=True) | Q(is_staff=True))
        return data


class AllProductListAPIView(generics.ListAPIView):
    serializer_class = ProductSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return Product.objects.filter(classification__in=[1, 2, 3, 5], status=True, delete=False).order_by('-id')


class CustomerAPIView(viewsets.ModelViewSet):
    queryset = Customer.objects.filter(delete=False, status=True)
    serializer_class = CustomerSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = Customer.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'customer deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class CreditSaleCustomerAPIView(viewsets.ModelViewSet):
    queryset = CreditSaleCustomer.objects.filter(delete=False, status=True)
    serializer_class = CreditSaleCustomerSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return CreditSaleCustomer.objects.filter(branch=self.request.user.branch, delete=False, status=True)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = CreditSaleCustomer.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'credit sale customer deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class CreditSalesAPIView(viewsets.ModelViewSet):
    queryset = CreditSales.objects.filter(delete=False, status=True)
    serializer_class = CreditSalesSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = CreditSales.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'credit sale deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class CreditSalesListAPIView(generics.ListAPIView):
    serializer_class = CreditSalesSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return CreditSales.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)


class CreditSettlementAPIView(viewsets.ModelViewSet):
    queryset = CreditSettlement.objects.filter(delete=False, status=True)
    serializer_class = CreditSettlementSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = CreditSettlement.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'credit settlement deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class CreditSettlementListAPIView(generics.ListAPIView):
    serializer_class = CreditSettlementSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return CreditSettlement.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)


class PettyCashAPIView(viewsets.ModelViewSet):
    queryset = PettyCash.objects.filter(delete=False, status=True)
    serializer_class = PettyCashSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = PettyCash.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'petty cash deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class PettyCashListAPIView(generics.RetrieveAPIView):
    serializer_class = PettyCashSerializer
    permission_classes = (IsAuthenticated,)

    def get_object(self):
        return PettyCash.objects.get(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)


class PettyCashRemarkDeleteAPIView(generics.DestroyAPIView):
    serializer_class = PettyCashSerializer
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = PettyCashRemark.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'petty cash remark deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class PettyCashCreateAPIView(generics.CreateAPIView):
    serializer_class = PettyCashSerializer
    permission_classes = (IsAuthenticated,)

    def get_serializer_context(self):
        return {
            'request': self.request,
            'format': self.format_kwarg,
            'view': self,
        }


class PettyCashPreviousListAPIView(generics.RetrieveAPIView):
    serializer_class = PettyCashListSerializer
    permission_classes = (IsAuthenticated,)

    def get_object(self):
        today = datetime.datetime.strptime(self.kwargs['date'], '%Y-%m-%d')
        yesterday = today - datetime.timedelta(days=1)
        data = PettyCash.objects.get(branch=self.request.user.branch, date__date=yesterday.date(), delete=False, status=True)
        return data


class SalesCountCreateAPIView(generics.CreateAPIView):
    serializer_class = SalesCountCreateSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request):
        for salescount_data in request.data:
            serializer = self.serializer_class(data=salescount_data, context={'branch': self.request.user.branch.pk})
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return Response({'message': 'Data Saved!'})


class SalesCountListAPIView(generics.ListAPIView):
    serializer_class = SalesCountlistSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        if self.kwargs['branch_id']==0:
            query = SalesCount.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)
        else:
            print(self.kwargs['branch_id'])
            query = SalesCount.objects.filter(branch=self.kwargs['branch_id'], date__date=self.kwargs['date'], delete=False, status=True)
        return query


class SalesCountDeleteAPIView(generics.DestroyAPIView):
    permission_classes = (IsAuthenticated,)

    def destroy(self, request, *args, **kwargs):
        destroy = SalesCount.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'sales count deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class BankCashReceivedDetailsAPIView(viewsets.ModelViewSet):
    queryset = BankCashReceivedDetails.objects.filter(delete=False, status=True)
    serializer_class = BankCashReceivedDetailsSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = BankCashReceivedDetails.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'bank cash received details deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class BankCashReceivedDetailsListAPIView(generics.ListAPIView):
    serializer_class = BankCashReceivedDetailsSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return BankCashReceivedDetails.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)


class DenominationAPIView(viewsets.ModelViewSet):
    queryset = Denomination.objects.filter(delete=False, status=True)
    serializer_class = DenominationSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request):
        for denomination_data in request.data:
            serializer = self.serializer_class(data=denomination_data, context={'branch': self.request.user.branch.pk})
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return Response({'message': 'Data Saved!'})

    def destroy(self, request, *args, **kwargs):
        destroy = Denomination.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'denomination deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class DenominationListAPIView(generics.ListAPIView):
    serializer_class = DenominationSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, date):
        query = Denomination.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)
        total = query.aggregate(overall_amount=Coalesce(Sum('total'), V(0)))
        return Response({
            "data": DenominationSerializer(query, many=True).data,
            "total": total['overall_amount']
        })


class DenominationUpdateAPIView(generics.UpdateAPIView):
    queryset = Denomination.objects.filter(delete=False)
    serializer_class = DenominationUpdateSerializer
    permission_classes = (IsAuthenticated,)

    def partial_update(self, request):
        for denomination_data in request.data:
            query = Denomination.objects.get(pk=denomination_data['id'])
            query.amount = denomination_data['amount']
            query.quantity = denomination_data['quantity']
            query.save()
        return Response({'message': 'Updated Successfully'}, status=status.HTTP_202_ACCEPTED)


class BranchCashManagementAPIView(viewsets.ModelViewSet):
    queryset = BranchCashManagement.objects.filter(delete=False, status=True)
    serializer_class = BranchCashManagementSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        today = datetime.date.today()
        date = today - datetime.timedelta(days=1)
        try:
            closing_cash = BranchCashManagement.objects.get(date__date=date, delete=False, status=True).closing_cash
        except BranchCashManagement.DoesNotExist:
            closing_cash = 0
        serializer.save(branch=self.request.user.branch, opening_cash=closing_cash)

    def destroy(self, request, *args, **kwargs):
        destroy = BranchCashManagement.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'denomination deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class BranchCashManagementListAPIView(generics.RetrieveAPIView):
    serializer_class = BranchCashManagementSerializer
    permission_classes = (IsAuthenticated,)

    def retrieve(self, request, *args, **kwargs):
        try:
            queryset = BranchCashManagement.objects.get(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)
        except BranchCashManagement.DoesNotExist:
            previous_day_data = BranchCashManagement.objects.filter(branch=self.request.user.branch, delete=False, status=True).aggregate(total_closing_cash=Coalesce(Sum('closing_cash'), V(0)))
            credit_sales_data = CreditSales.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True).aggregate(total_amount=Coalesce(Sum('amount'), V(0)))
            bank_cash_data = BankCashReceivedDetails.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True).aggregate(total_amount=Coalesce(Sum('amount'), V(0)))
            queryset = BranchCashManagement(branch=self.request.user.branch, date=self.kwargs['date'], opening_cash=previous_day_data['total_closing_cash'], credit_sales=credit_sales_data['total_amount'], bank_cash=bank_cash_data['total_amount'])
        return Response(self.serializer_class(queryset).data)


class CashHandoverDetailsAPIView(viewsets.ModelViewSet):
    queryset = CashHandover.objects.filter(delete=False, status=True)
    serializer_class = CashHandoverDetailsSerializer
    permission_classes = (IsAuthenticated,)

    def perform_create(self, serializer):
        serializer.save(branch=self.request.user.branch)

    def destroy(self, request, *args, **kwargs):
        destroy = CashHandover.objects.filter(pk=kwargs['pk']).update(status=False, delete=True)
        return Response({'message': 'cash handover details deleted successfully'}, status=status.HTTP_204_NO_CONTENT)


class CashHandoverDetailsListAPIView(generics.ListAPIView):
    serializer_class = CashHandoverDetailsSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return CashHandover.objects.filter(branch=self.request.user.branch, date__date=self.kwargs['date'], delete=False, status=True)


class UserProfileAPIView(generics.RetrieveAPIView):
    serializer_class = UserProfileSerializer
    permission_classes = (IsAuthenticated,)

    def get_object(self):
        return BaseUser.objects.get(pk=self.request.user.pk)


class BranchSpecificBillAPIView(generics.ListAPIView):
    serializer_class = BranchSpecificBillSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, *args, **kwargs):
        id = self.kwargs['id']
        date = self.kwargs['date']
        branch = self.kwargs['branch']

        if id == 1:
            query = WrongBill.objects.filter(date__date=date, branch=branch, delete=False, status=True)
            serializer_data = WrongBillSerializer(query, many=True)
        elif id == 2:
            query = FreeBill.objects.filter(date__date=date, branch=branch, delete=False, status=True)
            serializer_data = FreeBillSerializer(query, many=True)
        elif id == 3:
            query = EBMeter.objects.filter(branch=branch, delete=False, status=True)
            serializer_data = ElectricMeterElectricBillSerializer(query, many=True, context={'date': date})
        else:
            return Response({'message': 'No data found'})
        return Response(serializer_data.data, status=status.HTTP_200_OK)


class ElectricBillCreate(generics.CreateAPIView):
    serializer_class = ElectricBillSerializer
    permission_classes = (IsAuthenticated,)


    def create(self, request, date):
        for eb_data in request.data:
            serializer = self.serializer_class(data=eb_data, context={'branch': self.request.user.branch.pk, 'date': date})
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return Response({'message': 'Data Saved!'})

def my_cron_job():
    today = datetime.date.today()
    date = today - datetime.timedelta(days=1)

    salary_data = UserSalary.objects.filter(status=True, delete=False).values_list('user__pk', flat=True).distinct()
    attendance_data = UserAttendance.objects.filter(date=date, status=True, delete=False).values_list('user__pk', flat=True).distinct()
    user = BaseUser.objects.filter(pk__in=salary_data, is_superuser=False, is_active=True)
    abscent_users = user.exclude(pk__in=attendance_data)
    for user in abscent_users:
        data = UserAttendance.objects.create(user=user, abscent=True, date=date)


class CashDetailsAPIView(generics.ListAPIView):
    serializer_class = CashDetailsSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, *args, **kwargs):
        id = self.kwargs['id']
        date = self.kwargs['date']
        branch = self.kwargs['branch']

        if id == 1:
            try:
                query = PettyCash.objects.get(branch=branch, date__date=date, delete=False, status=True)
                serializer_data = PettyCashSerializer(query)
            except PettyCash.DoesNotExist:
                return Response([], status=status.HTTP_200_OK)
        elif id == 2:
            query = CreditSales.objects.filter(branch=branch, date__date=date, delete=False, status=True)
            serializer_data = CreditSalesSerializer(query, many=True)
        elif id == 3:
            query = CreditSettlement.objects.filter(branch=branch, date__date=date, delete=False, status=True)
            serializer_data = CreditSettlementSerializer(query, many=True)
        elif id == 4:
            query = Denomination.objects.filter(branch=branch, date__date=date, delete=False, status=True)
            total = query.aggregate(overall_amount=Coalesce(Sum('total'), V(0)))
            return Response({
                "data": DenominationSerializer(query, many=True).data,
                "total": total['overall_amount']
            })
        elif id == 5:
            query = BankCashReceivedDetails.objects.filter(branch=branch, date__date=date, delete=False, status=True)
            serializer_data = BankCashReceivedDetailsSerializer(query, many=True)
        elif id == 6:
            query = CashHandover.objects.filter(branch=branch, date__date=date, delete=False, status=True)
            serializer_data = CashHandoverDetailsSerializer(query, many=True)
        elif id == 7:
            try:
                query = BranchCashManagement.objects.get(branch=branch, date__date=date, delete=False, status=True)
                serializer_data = BranchCashManagementSerializer(query)
            except BranchCashManagement.DoesNotExist:
                return Response([], status=status.HTTP_200_OK)
        else:
            return Response({'message': 'No data found'})
        return Response(serializer_data.data, status=status.HTTP_200_OK)


class ProductStockInList(generics.ListAPIView):
    serializer_class = ProductStockInSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, classification):
        query = ProductBranchMapping.objects.get(branch=self.request.user.branch).product.order_by('-id')
        products = query.filter(classification__code=classification).order_by('-id')
        return Response(ProductStockInSerializer(products, context={'branch': self.request.user.branch.pk}, many=True).data)


class ProductPricingBatchCreateAPIView(generics.CreateAPIView):
    queryset = ProductPricingBatch.objects.filter(delete=False, status=True)
    serializer_class = ProductPricingBatchCreateSerializer
    permission_classes = (IsAuthenticated,)

    def create(self, request):
        for inventory_data in request.data:
            serializer = self.serializer_class(data=inventory_data, context={'branch': self.request.user.branch.pk})
            serializer.is_valid(raise_exception=True)
            serializer.save()
        return Response({'message': 'Data Saved!'})


class BranchSpecificFoodWastageListAPIView(generics.ListAPIView):
    serializer_class = FoodWastageSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return FoodWastage.objects.filter(date__date=self.kwargs['date'], branch=self.kwargs['branch'], delete=False, status=True)


class BranchSpecificOilConsumptionListAPIView(generics.ListAPIView):
    serializer_class = OilConsumptionSerializer
    permission_classes = (IsAuthenticated,)

    def get_queryset(self):
        return OilConsumption.objects.filter(branch=self.kwargs['branch'], date__date=self.kwargs['date'], delete=False, status=True).order_by('-id')


def ProductInventoryControlListToExcel(request, branch, date):

    wb = Workbook()
    ws = wb.active
    ws.title = "MyTestSheet"

    row_num = 1
    product_catagory = ['','','Operational product', 'Raw materials', 'Vegetable']
    for classification in range(2,5):

        querys = ProductBranchMapping.objects.get(branch=branch).product.order_by('-id')
        products = querys.filter(classification__code=classification).order_by('-id')

        columns = ['Date', 'Branch', 'Product / Unit', 'Openning Stock', 'Received Stock', 'Closing Stock', 'Usage']

        catagory_title = ws.cell(row=row_num, column=4, value=product_catagory[classification])
        ws.row_dimensions[row_num].height = 25
        catagory_title.font = Font(name='Calibri', size=13, bold=True, color='00695C')
        row_num += 1
        col_num = 0
        for value in columns:
            col_num = col_num+1
            cell = ws.cell(row=row_num, column=col_num, value = value)
            cell.font = Font(name='Chandas', bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
            column_letter = get_column_letter(col_num)
            column_dimensions = ws.column_dimensions[column_letter]
            column_dimensions.width = 20
            ws.row_dimensions[row_num].height = 25
            cell.fill = PatternFill(start_color='00838F',
                   end_color='00838F',
                   fill_type='solid')

        product_data = ProductInventoryControlSerializer(products, context={'branch': branch, 'date': date}, many=True).data

        for query in product_data:
            row_num += 1
            col_num = 0

            row = [date, Branch.objects.get(pk=branch).name, Product.objects.get(pk=query['product']).name+" - "+Unit.objects.get(pk=query['unit']).name, query['opening_stock'], query['received_stock'], query['closing_stock'], query['usage']]

            for value in row:
                col_num = col_num + 1
                value_cell = ws.cell(row=row_num, column=col_num, value = value)
                value_cell.alignment = Alignment(horizontal='center')
                value_cell.fill = PatternFill(start_color='E3F2FD',
                                    end_color='E3F2FD',
                                    fill_type='solid')
                if value == 0:
                    value_cell.font = Font(color='FF5252')

        row_num += 1
        ws['A'+str(row_num)].value = ' '
        row_num += 1
        ws.cell(row=row_num, column=1, value=" ")
        row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename=product_list.xlsx'
    file = wb.save(response)
    return response



class SlickPosProducts(generics.ListAPIView):
    serializer_class = BranchSpecificBillSerializer
    permission_classes = (IsAuthenticated,)

    def list(self, request, *args, **kwargs):
        products = requests.get('https://api.slickpos.com/api/product/list?accountId=7370d786-e7f5-436f-b608-0a7644fdb8e7', headers=headers)
        pos_products = products.json()
        for product in pos_products:
            slick_products = SlickposProducts.objects.create(
                slickpos_id=product['id'],
                name=product['name'],
                category_id=product['categoryId'],
                taxgroup_id=product['taxGroupId'],
                marked_price=product['markedPrice'],
                register_id=product['registerId'],
                variant_group_id=product['variantGroupIds'],
                addon_group_id=product['addonGroupId'],
                order_id=product['order']
            )

            if product['categoryId'] == '27033207-2484-42b4-9a90-808964af7230':
                department = 2
            elif product['categoryId'] == '11a0e6a3-ddaa-4648-b33e-fc0971464083':
                department = 5
            elif product['categoryId'] == '8ffaf637-e485-4328-ac30-c89a586622d9':
                department = 3
            elif product['categoryId'] == 'f3c6cb8e-30a2-41ba-8d7f-b3af611e5fe0':
                department = 6
            elif product['categoryId'] == '48feb8c8-2199-411d-82e3-20bf79908bfd':
                department = 1
            else:
                department = None
            all_products = Product.objects.create(
                product_id = product['id'],
                name = product['name'],
                classification = BranchProductClassification.objects.get(code=1),
                department =  BranchProductDepartment.objects.get(code=department)
            )

        world_json = os.path.join(os.getcwd(), 'countries+states+cities.json')
        with open(world_json) as f:
            world_data = json.load(f)
        for country in world_data:
            print(country)
            if country['id'] == 101:  # (for India) if you need all country dount use this
                country_data = Country.objects.create(
                    id=country['id'],
                    name=country['name'],
                    iso3=country['iso3'],
                    iso2=country['iso2'],
                    phone_code=country['phone_code'],
                    capital=country['capital'],
                    currency=country['currency']
                )
                for state in country['states']:
                    state_data = State.objects.create(
                        country=Country.objects.get(pk=country['id']),
                        id=state['id'],
                        name=state['name'],
                        state_code=state['state_code']
                    )
                    for city in state['cities']:
                        print(city['name'])
                        city_data = City.objects.create(
                            state=State.objects.get(pk=state['id']),
                            id=city['id'],
                            name=city['name'],
                            latitude=city['latitude'],
                            longitude=city['longitude']
                        )

        return Response({'message': 'Data Saved!'})


